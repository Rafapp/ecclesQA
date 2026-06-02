import os
import sys
import glob
import csv
import re
import openpyxl
from openpyxl.styles import PatternFill

import magic_runner as runner

# ── Constants ────────────────────────────────────────────────────────────────

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "mha_competencies_template.xlsx")

COMPETENCY_COL = {
    "SCT 1": "C", "SCT 2": "D", "SCT 3": "E",
    "RMT 1": "F", "RMT 2": "G", "RMT 3": "H",
    "LE 1":  "I", "LE 2":  "J", "LE 3":  "K", "LE 4": "L", "LE 5": "M",
    "BKS 1": "N", "BKS 2": "O", "BKS 3": "P", "BKS 4": "Q", "BKS 5": "R", "BKS 6": "S",
    "C 1":   "T", "C 2":   "U", "C 3":   "V", "C 4":   "W",
}
ALL_COMPETENCY_COLS = list(COMPETENCY_COL.values())

GRAY_FILL = PatternFill(patternType="solid", fgColor="D9D9D9")
NO_FILL   = PatternFill(patternType="none")

# ── CSV parsing ───────────────────────────────────────────────────────────────

def parse_course_code(filename):
    base = os.path.basename(filename)
    match = re.search(r'([A-Z]+)[_\s](\d{4})', base)
    return f"{match.group(1)} {match.group(2)}" if match else None

def parse_semester(filename):
    base = os.path.basename(filename)
    match = re.search(r'(Spring|Summer|Fall|Winter)[_\s](\d{4})', base, re.IGNORECASE)
    return f"{match.group(1).capitalize()} {match.group(2)}" if match else "Unknown"

def parse_competency_key(header):
    if not header.endswith(" result"):
        return None
    match = re.search(r'\(([^)]+)\)', header)
    return match.group(1).strip() if match else None

def load_csv(path):
    students = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("Student name", "").strip()
            if not name or name.lower() == "test student":
                continue
            sid   = row.get("Student ID", "").strip()
            sis   = row.get("Student SIS ID", "").strip()
            scores = {}
            for header, value in row.items():
                key = parse_competency_key(header)
                if key is None:
                    continue
                try:
                    scores[key] = float(value) if value.strip() else None
                except ValueError:
                    scores[key] = None
            students.append({"name": name, "student_id": sid, "sis_id": sis, "scores": scores})
    return students

# ── Template helpers ──────────────────────────────────────────────────────────

def parse_course_code_from_label(label):
    match = re.search(r'([A-Z]+)\s+(\d{4})', label)
    return f"{match.group(1)} {match.group(2)}" if match else None

def find_template_course_rows(ws):
    rows = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if cell.value:
            code = parse_course_code_from_label(str(cell.value))
            if code:
                rows[code] = cell.row
    return rows

def copy_template(template_wb):
    import io
    buf = io.BytesIO()
    template_wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf)

def apply_gray_row(ws, row_num):
    for col_letter in ALL_COMPETENCY_COLS:
        cell = ws[f"{col_letter}{row_num}"]
        cell.fill = GRAY_FILL
        cell.value = None

def write_score(ws, row_num, competency_key, value):
    col = COMPETENCY_COL.get(competency_key)
    if col is None:
        return
    cell = ws[f"{col}{row_num}"]
    cell.fill = NO_FILL
    cell.value = value

def sheet_name_for(student_name, student_id):
    safe_name = re.sub(r'[\\/*?:\[\]]', '', student_name)
    return f"{safe_name} {student_id}"[:31]

def copy_sheet_contents(src_ws, dst_ws):
    from copy import copy
    from openpyxl.cell.cell import MergedCell

    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
    for merge in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge))
    for row in src_ws.iter_rows():
        for src_cell in row:
            if isinstance(src_cell, MergedCell):
                continue
            dst_cell = dst_ws[src_cell.coordinate]
            if isinstance(dst_cell, MergedCell):
                continue
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font      = copy(src_cell.font)
                dst_cell.fill      = copy(src_cell.fill)
                dst_cell.border    = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 4:
        runner.run_error("Usage: mha_competencies.py <input_folder> <output_folder> <output_name>")

    input_folder  = sys.argv[1]
    output_folder = sys.argv[2]
    output_name   = sys.argv[3]

    # ── Step 1: Scan ──────────────────────────────────────────────────────────
    runner.step_start("scan", "Scanning for CSV files")

    if not os.path.isdir(input_folder):
        runner.step_error("scan", f"Input folder not found: {input_folder}")

    csv_files = sorted(glob.glob(os.path.join(input_folder, "*.csv")))
    if not csv_files:
        runner.step_error("scan", f"No .csv files found in: {input_folder}")

    if not os.path.exists(TEMPLATE_PATH):
        runner.step_error("scan", f"Template not found: {TEMPLATE_PATH}")

    runner.step_info(
        "scan",
        f"Found {len(csv_files)} CSV file(s) to process:",
        items=[os.path.basename(f) for f in csv_files],
        confirm=True,
    )
    runner.step_done("scan")

    # ── Step 2: Parse ─────────────────────────────────────────────────────────
    runner.step_start("parse", "Parsing student records")

    students_data  = {}
    semesters_seen = set()
    skipped_files  = []

    for csv_path in csv_files:
        course_code = parse_course_code(csv_path)
        semester    = parse_semester(csv_path)
        semesters_seen.add(semester)

        if not course_code:
            skipped_files.append(os.path.basename(csv_path))
            continue

        rows = load_csv(csv_path)
        for student in rows:
            key = student["name"].lower().strip()
            if key not in students_data:
                students_data[key] = {
                    "name":       student["name"],
                    "student_id": student["student_id"],
                    "sis_id":     student["sis_id"],
                    "courses":    {},
                }
            students_data[key]["courses"][course_code] = student["scores"]

    if skipped_files:
        runner.step_info("parse", "Skipped (could not parse course code):", items=skipped_files)

    if not students_data:
        runner.step_error("parse", "No student data found in any CSV file.")

    runner.step_info("parse", f"Parsed {len(students_data)} unique student(s) across {len(csv_files) - len(skipped_files)} course(s).")
    runner.step_done("parse")

    # ── Step 3: Build ─────────────────────────────────────────────────────────
    runner.step_start("build", "Building student worksheets")

    cohort_label = ", ".join(sorted(semesters_seen)) if semesters_seen else "Unknown"
    template_wb  = openpyxl.load_workbook(TEMPLATE_PATH)
    template_ws  = template_wb.active
    template_course_rows = find_template_course_rows(template_ws)

    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)

    sheet_names = []
    for _, student in sorted(students_data.items(), key=lambda x: x[1]["name"]):
        wb_copy = copy_template(template_wb)
        ws = wb_copy.active

        ws["B1"] = cohort_label
        ws["E1"] = student["name"]
        ws["G1"] = student["sis_id"] or student["student_id"]

        for course_code, row_num in template_course_rows.items():
            if course_code in student["courses"]:
                for comp_key, value in student["courses"][course_code].items():
                    write_score(ws, row_num, comp_key, value)
            else:
                apply_gray_row(ws, row_num)

        sname = sheet_name_for(student["name"], student["sis_id"] or student["student_id"])
        target_ws = output_wb.create_sheet(title=sname)
        copy_sheet_contents(ws, target_ws)
        sheet_names.append(sname)

    runner.step_info("build", f"Built {len(sheet_names)} sheet(s):", items=sheet_names)
    runner.step_done("build")

    # ── Step 4: Save ──────────────────────────────────────────────────────────
    runner.step_start("save", "Saving output workbook")

    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, f"{output_name}.xlsx")
    output_wb.save(output_path)

    runner.step_info("save", f"Saved to: {output_path}")
    runner.step_done("save")

    runner.run_done(f"Completed. {len(sheet_names)} student sheets saved to {output_path}")


if __name__ == "__main__":
    main()
